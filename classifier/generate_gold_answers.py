# -*- coding: utf-8 -*-
"""Generate gold labels for the local bid-classification spreadsheet.

This script is intentionally independent from the web/backend app. It reads a
two-column Excel dataset, calls two flagship models first, calls a third model
only when the first two disagree, and writes both an auditable JSONL checkpoint
and a final Excel workbook.
"""

from __future__ import annotations

import argparse
import concurrent.futures
import json
import os
import re
import sys
import time
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = PROJECT_ROOT / "classifier" / "bids.xlsx"
DEFAULT_OUTPUT_XLSX = PROJECT_ROOT / "classifier" / "bids_gold_answers.xlsx"
DEFAULT_OUTPUT_JSONL = PROJECT_ROOT / "classifier" / "bids_gold_answers.jsonl"

BAILIAN_BASE_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1"
DEEPSEEK_BASE_URL = "https://api.deepseek.com"

SYSTEM_PROMPT = """你是招投标信息行业分类助手。

你需要根据用户提供的“招标标题”和“发布企业”，从系统内置的“父级/子集”分类表中选择唯一一个最合适的父级和子集组合。

用户输入通常包含：

* 招标标题
* 发布企业

重要规则：

1. 分类表是固定分类表：每个“父级”下面有一组允许选择的“子集”。
2. 最终分类单位是“子集”，但输出时必须同时输出它对应的父级。
3. 不允许输出多个候选子集。
4. 不允许创造分类，父级和子集必须来自固定分类表。
5. 父级和子集必须存在对应关系：子集必须属于输出父级下的子集列表。
6. 不要做关键词、正则或词条命中的机械判断，必须根据语义、项目场景、业主/使用方和上下文综合判断。
7. 如果无法根据标题和发布企业稳定判断唯一子集，则输出可信=false，行业和细分行业为空字符串。

固定分类表如下：

父级【数据中心】允许选择的子集：
数据中心、智算中心、超算中心、数字科技、智能大数据、云计算、东数西算、算力中心。

父级【三大运营商】允许选择的子集：
移动、电信、联通。

父级【中企外投】允许选择的子集：
中企外投、海外储能、海外电解铝项目。

父级【熔盐储能】允许选择的子集：
熔盐储能、熔盐相关、热电联产、盐穴储能、盐穴相关。

父级【氢能制氢】允许选择的子集：
氢能制氢。

父级【油变专题】允许选择的子集：
油变挖掘、芳烃、炼油。

父级【煤炭行业】允许选择的子集：
煤炭行业、煤化工、燃煤发电。

父级【电子半导体】允许选择的子集：
半导体、中国电子、电力电子。

父级【航空航天】允许选择的子集：
机场、飞机制造、航空航天。

父级【石化行业】允许选择的子集：
中石化、中石油、中海油。

父级【电解铝】允许选择的子集：
电解铝、电解铜、焙阳极、碳素、钢厂/电炉类、烧碱化工。

父级【有色金属/冶炼】允许选择的子集：
磷矿、锂矿、其他矿类。

父级【常规】允许选择的子集：
交通行业、造纸行业、飞轮储能、抽水蓄能、风光储基地、空气压缩储能。

父级【其他】允许选择的子集：
变电站、产业园-厂房-基地、核电行业、医院行业、电池行业、风电发电、光伏项目、储能电站、火力发电、水利行业、化工行业、汽车制造、酒厂、上电旗下、哈电旗下、东电旗下、华润电力旗下、中盐旗下、隧道行业。

输出字段要求：

你必须只输出 JSON，不要输出 Markdown，不要输出解释性文字。

JSON 字段如下：

{
  "可信": true,
  "行业": "",
  "细分行业": "",
  "置信度": 0.0,
  "判断依据": []
}

字段含义：

1. “可信”：布尔值。能稳定判断唯一子集时为 true；无法稳定判断时为 false。
2. “行业”：输出父级名称。可信=false 时输出空字符串。
3. “细分行业”：输出子集名称。可信=false 时输出空字符串。
4. “置信度”：输出 0 到 1 之间的浮点数，表示分类置信度。
5. “判断依据”：输出字符串数组，简要说明判断依据。可信=false 时说明无法稳定判断的原因。

置信度规则：

1. 如果置信度低于 0.75，必须输出可信=false，行业和细分行业为空字符串。
2. 如果无法在固定分类表中找到唯一稳定匹配的子集，必须输出可信=false，行业和细分行业为空字符串。
3. 即使标题中出现了某些关键词，只要无法稳定判断项目场景、业主行业或使用方行业，也不要强行分类。

分类优先级：

第一优先级：项目场景。
如果标题中出现明确项目类型、建设对象、工程场景，应优先按照项目场景分类。
例如：

* 智算中心项目、算力中心项目、超算中心项目，优先归数据中心相关子集。
* 天然气热电项目，优先归热电联产。
* 选煤厂、煤矿、矿井、煤业项目、煤炭采掘项目，优先归煤炭行业。
* 煤化工项目，优先归煤化工。
* 电解铝项目，优先归电解铝。
* 机场扩建、机场改造，优先归机场。
* 储能电站项目，优先归储能电站。
* 光伏项目，优先归光伏项目。
* 风电场、风电项目，优先归风电发电。
* 生活垃圾焚烧发电、垃圾焚烧发电、垃圾发电、生物质发电、余热发电、燃气发电、燃油发电、燃煤发电、热电厂或发电厂机组项目，优先归“其他 / 火力发电”；如果标题明确是天然气热电联产、热电联产项目，则优先归“熔盐储能 / 热电联产”。
* 硬质合金、钨制品、钨粉、碳化钨、粉末冶金、合金刀片、数控刀片、涂层刀片、合金材料或新材料加工制造项目，在固定分类表没有专门“合金/新材料/粉末冶金”子集时，归“其他 / 化工行业”，判断依据中说明这是材料制造/新材料加工的兜底归并。

第二优先级：业主或使用方行业。
如果项目场景不明确，但发布企业、业主或使用方明显属于某行业，则按业主/使用方行业分类。
例如：

* 中煤、阳泉煤业、晋能煤业、煤矿集团，优先归煤炭行业。
* 中国移动、中国电信、中国联通，分别归移动、电信、联通。
* 中石油、中石化、中海油，分别归中石油、中石化、中海油。
* 医院、人民医院、妇幼保健院，归医院行业。

第三优先级：产品或设备名称。
产品和设备名称通常只是采购对象，不能直接决定行业。
例如：

* 变压器、干式变压器、开关柜、配电装置、电缆、启动器、移动变电站、箱变、高压柜、低压柜等，通常不能作为主分类依据。
* 如果标题中同时存在项目场景或业主行业，应忽略这些设备名对主分类的干扰。
* 只有当标题没有明确项目场景，也没有明确业主/使用方行业，只剩“变电站设施更换、变电站改造、变电站设备采购、公变专变、配电工程、配电房改造、电力增容及配电改造”这类工程对象时，才可归“其他 / 变电站”。

补充规则：

煤炭行业约束：

1. “矿用、矿山、选厂、采选工程、移动变电站”等词本身不能直接作为归入煤炭行业的依据。
2. 只有标题或发布企业中明确出现“煤、煤矿、煤业、焦煤、中煤、选煤厂、煤化工、采煤、采掘、矿井”等煤炭相关语义时，才可归入煤炭行业。
3. 如果标题中出现“金矿、锡、钼、铜、锂、磷矿、钛、有色、选矿厂、采选工程”等非煤矿业语义，不得因为出现“矿用、矿山、移动变电站”而归入煤炭行业。
4. 如果能稳定判断为锂矿、磷矿或其他非煤矿业项目，应优先归入固定分类表中对应的有色金属/冶炼相关子集。
5. 如果只能看出是泛矿业，但无法稳定判断具体矿种或对应子集，则输出可信=false。

充电桩与储能电站约束：

1. “充电桩、充电站、新能源汽车充电桩、高速服务区充电站”等项目，不等于储能电站。
2. 如果标题只体现充电桩或充电站建设，且没有出现“储能、电化学储能、储能系统、储能电站、BESS、电池舱、PCS、光储充”等储能相关语义，不得归入“储能电站”。
3. 充电桩、充电站、高速服务区充电站等项目，应优先考虑归入“常规 / 交通行业”。
4. 如果标题明确为“光储充一体化、储能充电站、含储能系统的充电站”等，才可结合项目主体判断是否归入储能电站。

光伏项目约束：

1. 只有标题中明确出现“光伏、PV、MWp、屋顶分布式光伏、地面光伏、光伏电站、光伏发电项目”等明确光伏语义时，才可归入“光伏项目”。
2. 仅出现“新能源、MW、箱变、箱式变压器、开关柜、框架采购”等信息，不足以稳定判断为光伏项目。
3. “新能源项目”可能是光伏、风电、风光储、储能或其他新能源类型，不得默认归入光伏项目。
4. 如果标题明确出现“风电、风电场、风力发电”，应优先归入“其他 / 风电发电”。
5. 如果标题明确出现“新能源基地、风光储基地、风光储一体化”等基地类综合新能源场景，应优先归入“常规 / 风光储基地”。
6. 如果只能看出是泛新能源，但无法稳定判断唯一子集，则输出可信=false。

火力发电与材料制造归并约束：

1. “生活垃圾焚烧发电、垃圾焚烧发电、垃圾发电、生物质发电、余热发电、燃气发电、燃油发电、燃煤发电、热电厂、发电厂机组”等均属于可稳定判断的发电项目场景，固定分类表中应归入“其他 / 火力发电”。
2. 不要因为固定分类表没有“垃圾发电、环保发电、生物质发电、余热发电”等完全同名子集而输出可信=false；这些场景应按火力发电的宽口径归并。
3. 如果标题明确是“热电联产”或“天然气热电联产”，按既有优先级归“熔盐储能 / 热电联产”，不要归火力发电。
4. “硬质合金、钨制品、钨粉、碳化钨、粉末冶金、合金刀片、数控刀片、涂层刀片、合金材料、新材料加工制造”等不是泛矿业，也不是“其他矿类”；在固定分类表缺少专门合金/新材料子集时，按“其他 / 化工行业”作为材料制造兜底分类。
5. 不要因为“硬质合金、合金刀片、涂层刀片”等没有完全同名子集而输出可信=false；只要项目场景明确为材料/合金/粉末冶金加工制造，应输出可信=true。

变电站与终端行业优先级：

1. “变电站、配电房、配电工程、公变专变、电力增容、输配电工程、站用变、主变压器”等词，很多时候只是项目中的电气配套工程，不能直接覆盖终端行业。
2. 如果标题中只有变电站、配电、增容、输配电等电气工程对象，且没有更明确的终端行业或项目场景，可以归入“其他 / 变电站”。
3. 如果标题中同时出现更明确的终端项目场景，例如光伏项目、风电场、储能电站、煤矿、煤化工、金矿、锂矿、磷矿、选厂、化工装置、核电厂、半导体厂、汽车制造项目、电池材料项目等，应优先按照终端项目场景分类，而不是被“变电站、配电柜、变压器、开关柜”等电气设备词干扰。
4. 判断时应区分“独立变电站工程”和“某行业项目中的变电站/配电配套工程”。前者可归变电站，后者应优先归终端行业。

特别注意：

1. “中国石油大学”是学校名称，不等于“中石油”，不能归石化行业 / 中石油。
2. “移动变电站”是设备名称，不等于“中国移动”，不能归三大运营商 / 移动。
3. “矿用移动变电站”中的“移动变电站”是设备名称，“矿用”也不能直接等于煤炭行业。
4. “热电项目”如果是项目场景，应优先看项目类型，而不是被“变压器、开关柜”等设备词干扰。
5. “采购设备”本身不是行业，必须看设备服务于什么项目或什么业主。
6. 父级“常规”和“其他”本身语义较弱，但仍必须严格遵守固定分类表中的父子对应关系。
7. 如果模型判断出的子集属于“常规”或“其他”下的子集，必须输出该子集在固定分类表中对应的正确父级。

输出示例：

{
  "可信": true,
  "行业": "其他",
  "细分行业": "光伏项目",
  "置信度": 0.92,
  "判断依据": ["标题明确出现光伏项目场景", "设备采购服务于光伏项目，按终端项目场景分类"]
}

无法稳定判断时输出示例：

{
  "可信": false,
  "行业": "",
  "细分行业": "",
  "置信度": 0.48,
  "判断依据": ["标题只体现通用设备采购，无法稳定判断项目场景或业主行业"]
}
"""

TAXONOMY: dict[str, set[str]] = {
    "数据中心": {"数据中心", "智算中心", "超算中心", "数字科技", "智能大数据", "云计算", "东数西算", "算力中心"},
    "三大运营商": {"移动", "电信", "联通"},
    "中企外投": {"中企外投", "海外储能", "海外电解铝项目"},
    "熔盐储能": {"熔盐储能", "熔盐相关", "热电联产", "盐穴储能", "盐穴相关"},
    "氢能制氢": {"氢能制氢"},
    "油变专题": {"油变挖掘", "芳烃", "炼油"},
    "煤炭行业": {"煤炭行业", "煤化工", "燃煤发电"},
    "电子半导体": {"半导体", "中国电子", "电力电子"},
    "航空航天": {"机场", "飞机制造", "航空航天"},
    "石化行业": {"中石化", "中石油", "中海油"},
    "电解铝": {"电解铝", "电解铜", "焙阳极", "碳素", "钢厂/电炉类", "烧碱化工"},
    "有色金属/冶炼": {"磷矿", "锂矿", "其他矿类"},
    "常规": {"交通行业", "造纸行业", "飞轮储能", "抽水蓄能", "风光储基地", "空气压缩储能"},
    "其他": {
        "变电站",
        "产业园-厂房-基地",
        "核电行业",
        "医院行业",
        "电池行业",
        "风电发电",
        "光伏项目",
        "储能电站",
        "火力发电",
        "水利行业",
        "化工行业",
        "汽车制造",
        "酒厂",
        "上电旗下",
        "哈电旗下",
        "东电旗下",
        "华润电力旗下",
        "中盐旗下",
        "隧道行业",
    },
}

EXPECTED_FIELDS = ["可信", "行业", "细分行业", "置信度", "判断依据"]


@dataclass(frozen=True)
class ModelSpec:
    key: str
    label: str
    provider: str
    base_url: str
    api_key_envs: tuple[str, ...]
    model: str


@dataclass
class ModelResult:
    model_key: str
    label: str
    ok: bool
    normalized_vote: tuple[bool, str, str] | None
    answer: dict[str, Any] | None
    raw_text: str
    error: str = ""
    attempts: int = 0
    duration_ms: int = 0


def load_env_files() -> None:
    protected_env_keys = {key for key, value in os.environ.items() if value}
    for path in (PROJECT_ROOT / ".env", PROJECT_ROOT / "backend" / ".env", PROJECT_ROOT / "classifier" / ".env"):
        if path.exists():
            load_simple_env_file(path, protected_env_keys)


def load_simple_env_file(path: Path, protected_env_keys: set[str]) -> None:
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        if not key or key in protected_env_keys:
            continue
        value = value.strip().strip('"').strip("'")
        os.environ[key] = value


def build_user_payload(title: str, company: str) -> str:
    return json.dumps({"招标标题": title, "发布企业": company}, ensure_ascii=False)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def call_chat_completion(
    spec: ModelSpec,
    title: str,
    company: str,
    temperature: float,
    top_p: float,
    max_tokens: int,
    timeout: float,
    retries: int,
    json_mode: bool,
) -> ModelResult:
    try:
        import httpx
    except ImportError as exc:  # pragma: no cover - dependency guard for local runs
        raise SystemExit(
            "Missing dependency: httpx. Install project requirements in your eval env, for example:\n"
            "  pip install -r backend/requirements.txt\n"
        ) from exc

    api_key = first_env(spec.api_key_envs)
    if not api_key:
        return ModelResult(
            model_key=spec.key,
            label=spec.label,
            ok=False,
            normalized_vote=None,
            answer=None,
            raw_text="",
            error=f"Missing API key env: {'/'.join(spec.api_key_envs)}",
        )

    payload: dict[str, Any] = {
        "model": spec.model,
        "temperature": temperature,
        "top_p": top_p,
        "max_tokens": max_tokens,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": build_user_payload(title, company)},
        ],
    }
    if json_mode:
        payload["response_format"] = {"type": "json_object"}

    url = spec.base_url.rstrip("/") + "/chat/completions"
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    started = time.perf_counter()
    last_error = ""
    attempts = 0

    for attempt in range(1, retries + 1):
        attempts = attempt
        try:
            with httpx.Client(timeout=timeout) as client:
                response = client.post(url, headers=headers, json=payload)
            if response.status_code >= 400 and json_mode and "response_format" in payload:
                # Some OpenAI-compatible endpoints reject JSON mode. Retry this
                # attempt without it before counting the provider as failed.
                fallback_payload = dict(payload)
                fallback_payload.pop("response_format", None)
                with httpx.Client(timeout=timeout) as client:
                    response = client.post(url, headers=headers, json=fallback_payload)
            response.raise_for_status()
            data = response.json()
            raw_text = extract_content(data)
            answer = normalize_answer(parse_json_object(raw_text))
            return ModelResult(
                model_key=spec.key,
                label=spec.label,
                ok=True,
                normalized_vote=vote_key(answer),
                answer=answer,
                raw_text=raw_text,
                attempts=attempts,
                duration_ms=int((time.perf_counter() - started) * 1000),
            )
        except Exception as exc:  # noqa: BLE001 - preserves provider error text for audit
            last_error = format_exception(exc)
            if attempt < retries:
                time.sleep(min(8, 0.8 * 2 ** (attempt - 1)))

    return ModelResult(
        model_key=spec.key,
        label=spec.label,
        ok=False,
        normalized_vote=None,
        answer=None,
        raw_text="",
        error=last_error,
        attempts=attempts,
        duration_ms=int((time.perf_counter() - started) * 1000),
    )


def first_env(names: tuple[str, ...]) -> str:
    for name in names:
        value = os.getenv(name, "").strip()
        if value:
            return value
    return ""


def extract_content(response_data: dict[str, Any]) -> str:
    choices = response_data.get("choices") or []
    if not choices:
        raise ValueError(f"No choices in response: {json.dumps(response_data, ensure_ascii=False)[:1000]}")
    message = choices[0].get("message") or {}
    content = message.get("content", "")
    if isinstance(content, str):
        return content.strip()
    if isinstance(content, list):
        parts = []
        for item in content:
            if isinstance(item, str):
                parts.append(item)
            elif isinstance(item, dict):
                parts.append(str(item.get("text") or item.get("content") or ""))
        return "".join(parts).strip()
    return str(content).strip()


def parse_json_object(text: str) -> dict[str, Any]:
    if not text:
        raise ValueError("Empty model output")
    cleaned = text.strip()
    fenced = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", cleaned, flags=re.DOTALL | re.IGNORECASE)
    if fenced:
        cleaned = fenced.group(1).strip()
    if not cleaned.startswith("{"):
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start >= 0 and end > start:
            cleaned = cleaned[start : end + 1]
    data = json.loads(cleaned)
    if not isinstance(data, dict):
        raise ValueError("Model output JSON is not an object")
    return data


def normalize_answer(data: dict[str, Any]) -> dict[str, Any]:
    answer = {field: data.get(field) for field in EXPECTED_FIELDS}
    trusted = bool(answer.get("可信"))
    industry = clean_text(answer.get("行业"))
    subindustry = clean_text(answer.get("细分行业"))
    try:
        confidence = float(answer.get("置信度") or 0)
    except (TypeError, ValueError):
        confidence = 0.0

    reasons = answer.get("判断依据")
    if isinstance(reasons, str):
        reasons = [reasons]
    elif not isinstance(reasons, list):
        reasons = []
    reasons = [clean_text(item) for item in reasons if clean_text(item)]

    invalid_reason = ""
    if trusted:
        if confidence < 0.75:
            invalid_reason = "模型置信度低于0.75，按规则改为不可信"
        elif industry not in TAXONOMY:
            invalid_reason = f"行业不在固定分类表中：{industry}"
        elif subindustry not in TAXONOMY.get(industry, set()):
            invalid_reason = f"细分行业不属于行业“{industry}”：{subindustry}"

    if invalid_reason:
        trusted = False
        industry = ""
        subindustry = ""
        reasons.append(invalid_reason)

    if not trusted:
        industry = ""
        subindustry = ""

    return {
        "可信": trusted,
        "行业": industry,
        "细分行业": subindustry,
        "置信度": max(0.0, min(1.0, confidence)),
        "判断依据": reasons or (["模型未提供可审计判断依据"] if trusted else ["无法稳定判断唯一子集"]),
    }


def vote_key(answer: dict[str, Any]) -> tuple[bool, str, str]:
    if not answer.get("可信"):
        return (False, "", "")
    return (True, answer.get("行业", ""), answer.get("细分行业", ""))


def choose_final_answer(results: list[ModelResult]) -> tuple[dict[str, Any], str, list[str]]:
    ok_results = [result for result in results if result.ok and result.normalized_vote is not None and result.answer]
    if not ok_results:
        return (
            {
                "可信": False,
                "行业": "",
                "细分行业": "",
                "置信度": 0.0,
                "判断依据": ["三个模型均未返回可解析的有效JSON"],
            },
            "all_failed",
            [],
        )

    counter = Counter(result.normalized_vote for result in ok_results)
    vote, count = counter.most_common(1)[0]
    voters = [result.label for result in ok_results if result.normalized_vote == vote]
    if count >= 2:
        selected = [result.answer for result in ok_results if result.normalized_vote == vote]
        avg_confidence = sum(float(item.get("置信度", 0) or 0) for item in selected) / len(selected)
        reasons: list[str] = []
        for item in selected:
            for reason in item.get("判断依据", []):
                if reason not in reasons:
                    reasons.append(reason)
        trusted, industry, subindustry = vote
        return (
            {
                "可信": trusted,
                "行业": industry if trusted else "",
                "细分行业": subindustry if trusted else "",
                "置信度": round(avg_confidence, 4),
                "判断依据": reasons,
            },
            "majority",
            voters,
        )

    labels = [format_model_vote(result) for result in results]
    return (
        {
            "可信": False,
            "行业": "",
            "细分行业": "",
            "置信度": 0.0,
            "判断依据": ["三个模型未形成多数一致结论：" + "；".join(labels)],
        },
        "no_majority",
        [],
    )


def first_wave_failure_answer(results: list[ModelResult]) -> dict[str, Any]:
    labels = [format_model_vote(result) for result in results]
    return {
        "可信": False,
        "行业": "",
        "细分行业": "",
        "置信度": 0.0,
        "判断依据": ["首轮模型未全部成功，未按三模型投票生成标准答案：" + "；".join(labels)],
    }


def format_model_vote(result: ModelResult) -> str:
    if not result.ok:
        return f"{result.label}=调用失败({result.error or 'unknown error'})"
    return f"{result.label}={format_vote(result.normalized_vote)}"


def format_vote(vote: tuple[bool, str, str] | None) -> str:
    if vote is None:
        return "无有效票"
    trusted, industry, subindustry = vote
    if not trusted:
        return "不可信"
    return f"{industry}/{subindustry}"


def read_input_rows(path: Path, title_col: str | None, company_col: str | None) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard for local runs
        raise SystemExit(
            "Missing dependency: openpyxl. Install project requirements in your eval env, for example:\n"
            "  pip install -r backend/requirements.txt\n"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [clean_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    if len(headers) < 2:
        raise ValueError("Input workbook must contain at least two columns with headers.")

    title_index = resolve_column_index(headers, title_col, default=0)
    company_index = resolve_column_index(headers, company_col, default=1)
    rows: list[dict[str, Any]] = []
    for excel_row_num, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        title = clean_text(cells[title_index] if title_index < len(cells) else "")
        company = clean_text(cells[company_index] if company_index < len(cells) else "")
        if not title and not company:
            continue
        rows.append({"excel_row": excel_row_num, "title": title, "company": company})
    workbook.close()
    return headers, rows


def resolve_column_index(headers: list[str], requested: str | None, default: int) -> int:
    if not requested:
        return default
    if requested in headers:
        return headers.index(requested)
    normalized = {header.strip().lower(): index for index, header in enumerate(headers)}
    key = requested.strip().lower()
    if key in normalized:
        return normalized[key]
    raise ValueError(f"Column not found: {requested}. Available headers: {headers}")


def load_existing_jsonl(path: Path) -> dict[int, dict[str, Any]]:
    if not path.exists():
        return {}
    records: dict[int, dict[str, Any]] = {}
    with path.open("r", encoding="utf-8") as file:
        for line in file:
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                continue
            excel_row = int(record.get("excel_row", 0) or 0)
            if excel_row:
                records[excel_row] = record
    return records


def append_jsonl(path: Path, record: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as file:
        file.write(json.dumps(record, ensure_ascii=False) + "\n")


def save_output_workbook(
    source_headers: list[str],
    source_rows: list[dict[str, Any]],
    records: dict[int, dict[str, Any]],
    output_path: Path,
) -> Path:
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - dependency guard for local runs
        raise SystemExit(
            "Missing dependency: openpyxl. Install project requirements in your eval env, for example:\n"
            "  pip install -r backend/requirements.txt\n"
        ) from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "gold_answers"
    output_headers = [
        "源Excel行号",
        source_headers[0] if source_headers else "招标标题",
        source_headers[1] if len(source_headers) > 1 else "发布企业",
        "标准_可信",
        "标准_行业",
        "标准_细分行业",
        "标准_置信度",
        "标准_判断依据",
        "一致状态",
        "多数模型",
        "deepseek_v4_pro_结果",
        "qwen3_7_max_结果",
        "glm5_2_结果",
        "错误信息",
    ]
    sheet.append(output_headers)

    for source in source_rows:
        record = records.get(source["excel_row"])
        if record:
            final_answer = record.get("final_answer") or {}
            model_results = record.get("model_results") or {}
            errors = [
                f"{key}: {value.get('error')}"
                for key, value in model_results.items()
                if isinstance(value, dict) and value.get("error")
            ]
            sheet.append(
                [
                    source["excel_row"],
                    source["title"],
                    source["company"],
                    final_answer.get("可信", ""),
                    final_answer.get("行业", ""),
                    final_answer.get("细分行业", ""),
                    final_answer.get("置信度", ""),
                    "；".join(final_answer.get("判断依据", [])),
                    record.get("agreement_status", ""),
                    "、".join(record.get("majority_models", [])),
                    json.dumps(model_results.get("deepseek_v4_pro", {}).get("answer"), ensure_ascii=False),
                    json.dumps(model_results.get("qwen3_7_max", {}).get("answer"), ensure_ascii=False),
                    json.dumps(model_results.get("glm5_2", {}).get("answer"), ensure_ascii=False)
                    if "glm5_2" in model_results
                    else "",
                    "；".join(errors),
                ]
            )
        else:
            sheet.append([source["excel_row"], source["title"], source["company"]])

    for column_cells in sheet.columns:
        max_length = max(len(clean_text(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 80)
    sheet.freeze_panes = "A2"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(output_path)
        return output_path
    except PermissionError:
        fallback_path = timestamped_output_path(output_path)
        workbook.save(fallback_path)
        print(
            f"Output workbook is locked or not writable: {output_path}. "
            f"Saved a fallback copy instead: {fallback_path}",
            file=sys.stderr,
        )
        return fallback_path


def timestamped_output_path(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return path.with_name(f"{path.stem}_{stamp}{path.suffix}")


def process_row(row: dict[str, Any], specs: dict[str, ModelSpec], args: argparse.Namespace) -> dict[str, Any]:
    first_wave_keys = ["deepseek_v4_pro", "qwen3_7_max"]
    first_results: list[ModelResult] = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        future_to_key = {
            executor.submit(
                call_chat_completion,
                specs[key],
                row["title"],
                row["company"],
                args.temperature,
                args.top_p,
                args.max_tokens,
                args.timeout,
                args.retries,
                not args.no_json_mode,
            ): key
            for key in first_wave_keys
        }
        for future in concurrent.futures.as_completed(future_to_key):
            first_results.append(future.result())

    all_results = sorted(first_results, key=lambda result: first_wave_keys.index(result.model_key))
    if len(first_results) != 2 or any(not result.ok for result in first_results):
        return {
            "excel_row": row["excel_row"],
            "title": row["title"],
            "company": row["company"],
            "final_answer": first_wave_failure_answer(all_results),
            "agreement_status": "first_wave_failed",
            "majority_models": [],
            "model_results": {result.model_key: model_result_to_dict(result) for result in all_results},
        }

    need_glm = first_results[0].normalized_vote != first_results[1].normalized_vote
    if need_glm:
        all_results.append(
            call_chat_completion(
                specs["glm5_2"],
                row["title"],
                row["company"],
                args.temperature,
                args.top_p,
                args.max_tokens,
                args.timeout,
                args.retries,
                not args.no_json_mode,
            )
        )

    final_answer, agreement_status, majority_models = choose_final_answer(all_results)
    return {
        "excel_row": row["excel_row"],
        "title": row["title"],
        "company": row["company"],
        "final_answer": final_answer,
        "agreement_status": "first_two_agree" if not need_glm and agreement_status == "majority" else agreement_status,
        "majority_models": majority_models,
        "model_results": {result.model_key: model_result_to_dict(result) for result in all_results},
    }


def model_result_to_dict(result: ModelResult) -> dict[str, Any]:
    return {
        "label": result.label,
        "ok": result.ok,
        "vote": list(result.normalized_vote) if result.normalized_vote is not None else None,
        "answer": result.answer,
        "raw_text": result.raw_text,
        "error": result.error,
        "attempts": result.attempts,
        "duration_ms": result.duration_ms,
    }


def format_exception(exc: Exception) -> str:
    message = str(exc)
    return f"{exc.__class__.__name__}: {message}" if message else exc.__class__.__name__


def build_model_specs(args: argparse.Namespace) -> dict[str, ModelSpec]:
    return {
        "deepseek_v4_pro": ModelSpec(
            key="deepseek_v4_pro",
            label="deepseek v4 pro",
            provider="deepseek",
            base_url=args.deepseek_base_url,
            api_key_envs=("DEEPSEEK_API_KEY",),
            model=args.deepseek_model,
        ),
        "qwen3_7_max": ModelSpec(
            key="qwen3_7_max",
            label="千问3.7max",
            provider="bailian",
            base_url=args.bailian_base_url,
            api_key_envs=("BAILIAN_API_KEY", "DASHSCOPE_API_KEY"),
            model=args.qwen_model,
        ),
        "glm5_2": ModelSpec(
            key="glm5_2",
            label="GLM5.2",
            provider="bailian",
            base_url=args.bailian_base_url,
            api_key_envs=("BAILIAN_API_KEY", "DASHSCOPE_API_KEY"),
            model=args.glm_model,
        ),
    }


def validate_required_api_keys(specs: dict[str, ModelSpec]) -> list[str]:
    missing: list[str] = []
    required_groups: list[tuple[str, tuple[str, ...]]] = [
        (specs["deepseek_v4_pro"].label, specs["deepseek_v4_pro"].api_key_envs),
        (specs["qwen3_7_max"].label, specs["qwen3_7_max"].api_key_envs),
        (specs["glm5_2"].label, specs["glm5_2"].api_key_envs),
    ]
    for label, envs in required_groups:
        if not first_env(envs):
            missing.append(f"{label}: {' or '.join(envs)}")
    return missing


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate majority-vote gold answers for classifier/bids.xlsx."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Input .xlsx path.")
    parser.add_argument("--output-xlsx", type=Path, default=DEFAULT_OUTPUT_XLSX, help="Final annotated .xlsx path.")
    parser.add_argument("--output-jsonl", type=Path, default=DEFAULT_OUTPUT_JSONL, help="Checkpoint/audit JSONL path.")
    parser.add_argument("--title-col", default=None, help="Title column header. Defaults to the first column.")
    parser.add_argument("--company-col", default=None, help="Company column header. Defaults to the second column.")
    parser.add_argument("--deepseek-base-url", default=os.getenv("DEEPSEEK_BASE_URL", DEEPSEEK_BASE_URL))
    parser.add_argument("--bailian-base-url", default=os.getenv("BAILIAN_BASE_URL", BAILIAN_BASE_URL))
    parser.add_argument("--deepseek-model", default=os.getenv("DEEPSEEK_GOLD_MODEL", "deepseek-v4-pro"))
    parser.add_argument("--qwen-model", default=os.getenv("QWEN_GOLD_MODEL", "qwen3.7-max"))
    parser.add_argument("--glm-model", default=os.getenv("GLM_GOLD_MODEL", "glm-5.2"))
    parser.add_argument("--temperature", type=float, default=float(os.getenv("GOLD_TEMPERATURE", "0")))
    parser.add_argument("--top-p", type=float, default=float(os.getenv("GOLD_TOP_P", "0.9")))
    parser.add_argument("--max-tokens", type=int, default=int(os.getenv("GOLD_MAX_TOKENS", "900")))
    parser.add_argument("--timeout", type=float, default=float(os.getenv("GOLD_TIMEOUT", "120")))
    parser.add_argument("--retries", type=int, default=int(os.getenv("GOLD_RETRIES", "3")))
    parser.add_argument("--limit", type=int, default=0, help="Only process the first N rows after filtering. 0 means all.")
    parser.add_argument("--start-row", type=int, default=2, help="First Excel row number to process.")
    parser.add_argument("--sleep", type=float, default=0.0, help="Seconds to sleep after each processed row.")
    parser.add_argument("--force", action="store_true", help="Ignore existing JSONL checkpoint and regenerate rows.")
    parser.add_argument("--no-json-mode", action="store_true", help="Do not request response_format=json_object.")
    parser.add_argument("--dry-run", action="store_true", help="Read inputs and print configuration without calling APIs.")
    return parser.parse_args()


def main() -> int:
    load_env_files()
    args = parse_args()
    specs = build_model_specs(args)
    headers, rows = read_input_rows(args.input, args.title_col, args.company_col)
    rows = [row for row in rows if row["excel_row"] >= args.start_row]
    if args.limit:
        rows = rows[: args.limit]

    existing = {} if args.force else load_existing_jsonl(args.output_jsonl)
    completed = dict(existing)
    pending = [row for row in rows if row["excel_row"] not in completed]

    print(f"Input: {args.input}")
    print(f"Rows selected: {len(rows)}; pending: {len(pending)}; checkpointed: {len(completed)}")
    print(
        "Models: "
        f"DeepSeek={specs['deepseek_v4_pro'].model}; "
        f"Qwen={specs['qwen3_7_max'].model}; "
        f"GLM={specs['glm5_2'].model}"
    )
    if args.dry_run:
        print("Dry run only. No API calls and no output files were written.")
        if rows:
            sample = rows[0]
            print(f"Sample row {sample['excel_row']}: title={sample['title']!r}; company={sample['company']!r}")
        return 0

    missing_keys = validate_required_api_keys(specs)
    if missing_keys:
        print("Missing required API key environment variables:", file=sys.stderr)
        for item in missing_keys:
            print(f"  - {item}", file=sys.stderr)
        print("No API calls were made. Set the keys and rerun, or put them in .env/backend\\.env.", file=sys.stderr)
        return 2

    if args.force and args.output_jsonl.exists():
        args.output_jsonl.write_text("", encoding="utf-8")
        completed = {}
        pending = rows

    for index, row in enumerate(pending, start=1):
        print(f"[{index}/{len(pending)}] Excel row {row['excel_row']}: {row['title'][:60]}")
        record = process_row(row, specs, args)
        append_jsonl(args.output_jsonl, record)
        completed[row["excel_row"]] = record
        final_answer = record["final_answer"]
        print(
            "  -> "
            f"{record['agreement_status']} | "
            f"{final_answer.get('可信')} | "
            f"{final_answer.get('行业')}/{final_answer.get('细分行业')} | "
            f"conf={final_answer.get('置信度')}"
        )
        if args.sleep:
            time.sleep(args.sleep)

    # Preserve already checkpointed rows outside the current selected range too.
    all_records = load_existing_jsonl(args.output_jsonl)
    all_records.update(completed)
    saved_xlsx = save_output_workbook(
        headers,
        read_input_rows(args.input, args.title_col, args.company_col)[1],
        all_records,
        args.output_xlsx,
    )
    print(f"Saved JSONL: {args.output_jsonl}")
    print(f"Saved XLSX: {saved_xlsx}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
